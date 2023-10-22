import requests
import json 
import openpyxl
import csv
import pandas as pd
from openpyxl import load_workbook

def get_microsoft_data():

    header = {
            "accept": "application/json",
            "Content-Type": "application/json"
            }
    payload = {}

    url = "https://app-omaha-prod.azurewebsites.net/api/PublishedListings/Export"

    try:

        # returns excel file
        resp = requests.get(url , headers=header , data=payload)


        # excel is binary use response content to write the data
        with open("microsoft_eol.xlsx" ,mode="wb" ) as file:
            file.write(resp.content)
        print("Data has been added to Excel file.. ")


        # copy old excel data into the new excel file
        # read the source Excel file into a pandas dataframe
        source_df = pd.read_excel("microsoft_eol.xlsx")

        # create a new Excel writer object for the destination file
        writer = pd.ExcelWriter("Updated_microsoft_eol.xlsx", engine='xlsxwriter')

        # write the source dataframe to the destination file
        source_df.to_excel(writer, sheet_name='Sheet1', index=False )

        # save the destination file
        writer.save()

       

        # remove the empty rows
        df = pd.read_excel("Updated_microsoft_eol.xlsx")

        # remove rows that are completely empty
        df = df.dropna(subset=['Unnamed: 0'])

        # save the updated dataframe to a new Excel file
        df.to_excel("Updated_microsoft_eol.xlsx", index=False)
        print("Removed blank rows")

        # remove the first row
        # open excel file 
#         wb = openpyxl.load_workbook("Updated_microsoft_eol.xlsx")
#         # create active sheet object
#         sheet1 = wb.active
#         # delete row number 3
#         sheet1.delete_rows(1)
#         sheet1.delete_rows(1)
# #         sheet1.delete_rows(1)
#         wb.save("Updated_microsoft_eol.xlsx")
#         wb.close()
#         print("Removed the first ")

        # add vendor column to excel file
        workbook = load_workbook(filename='Updated_microsoft_eol.xlsx')
        worksheet = workbook['Sheet1']
        column_range = worksheet['A'] # Example for column A, adjust as needed
        column_size = len(column_range)
        print(f"Column size: {column_size}")
        value = "microsoft"
        for i in range(1, column_size+1):
            worksheet.cell(row=i, column=12, value=value)

        worksheet.cell(row=1 , column=12 , value="vendor")
        workbook.save(filename='Updated_microsoft_eol.xlsx')
        print("Data has been added to Updated_microsoft_eol.xlsx file ")
            # save micorsoft xlsx data csv
                # read the xlsx file using pandas
        xlsx_file = pd.read_excel('Updated_microsoft_eol.xlsx')

        # save the data to a csv file
        xlsx_file.to_csv('microsoft_eol.csv', index=False)
        print("Updated_microsoft_eol.xlsx file converted into microsoft_eol.csv")
    except Exception as e:
        print("Could not get the data for microsoft")
        print("Error occured ", e)



